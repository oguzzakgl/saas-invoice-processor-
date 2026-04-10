import asyncio
import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from engine import Engine

# Windows'da asyncio'nun 'Could not contact DNS servers' veya SSL hataları vermesini
# önlemek için klasik Selector tipine (WindowsSelectorEventLoopPolicy) geçiriyoruz.
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())


# ─────────────────────────────────────────────────────────────────
# MODÜL 1: GELİŞMİŞ DENETİM (Cross-Check Validation)
# ─────────────────────────────────────────────────────────────────

def validate_financials(data) -> dict:
    """
    Belgeden çekilen finansal verilerin matematiksel tutarlılığını denetler.
    
    Kural: Ara Toplam (subtotal) + KDV (tax_amount) = Genel Toplam (total_amount)
    Tolerans: %1 yuvarlama payı uygulanır (örn: 1000.00 TL için ±10 TL tolere edilir).
    
    Parametreler:
        data: LLM'den dönen ExtractedData objesi (fatura_detaylari alanını içerir)
    
    Döndürdüğü Değer:
        dict: { "valid": bool, "message": str, "difference": float }
    """
    # Başarısız durum için varsayılan sonuç
    sonuc = {"valid": False, "message": "Veri yetersiz veya eksik.", "difference": 0.0}
    
    try:
        # Verinin fatura detayı içerip içermediğini kontrol et
        if not data or not getattr(data, 'fatura_detaylari', None):
            sonuc["message"] = "Belgede fatura detayı bulunamadı (sözleşme veya boş belge olabilir)."
            return sonuc
        
        inv = data.fatura_detaylari
        
        # Üç kritik alan da dolu mu? (None değer varsa hesap yapılamaz)
        if inv.ara_toplam is None or inv.kdv_tutari is None or inv.toplam_tutar is None:
            sonuc["message"] = "⚠️ Finansal alanlar eksik — Validasyon yapılamadı."
            return sonuc
        
        # Matematiksel denetim: Ara Toplam + KDV = Genel Toplam
        hesaplanan_toplam = inv.ara_toplam + inv.kdv_tutari
        fark = abs(hesaplanan_toplam - inv.toplam_tutar)
        
        # %1 tolerans payı: Toplam tutarın %1'i sınır olarak belirlendi
        tolerans = inv.toplam_tutar * 0.01
        
        sonuc["difference"] = round(fark, 2)
        
        if fark <= tolerans:
            # Denklem sağlandı → Finansal tutarlılık onaylandı
            sonuc["valid"] = True
            sonuc["message"] = f"✅ Finansal Tutarlılık Onaylandı (Fark: {fark:.2f}, Tolerans: {tolerans:.2f})"
        else:
            # Denklem sağlanamadı → Tutarsızlık tespit edildi
            sonuc["valid"] = False
            sonuc["message"] = (
                f"🚨 Finansal Tutarsızlık Tespit Edildi! "
                f"Beklenen: {hesaplanan_toplam:.2f}, Belgede: {inv.toplam_tutar:.2f}, "
                f"Fark: {fark:.2f} (İzin verilen tolerans: {tolerans:.2f})"
            )
            print(f"[UYARI] {sonuc['message']}")
    
    except (TypeError, AttributeError) as e:
        # Beklenmedik veri tipi veya eksik alan hatası → sistem çökmeden devam eder
        sonuc["message"] = f"❌ Validasyon sırasında veri hatası oluştu: {e}"
    except Exception as e:
        sonuc["message"] = f"❌ Beklenmedik validasyon hatası: {e}"
    
    return sonuc


# ─────────────────────────────────────────────────────────────────
# MODÜL 2: RAPORLAMA VE OTOMASYON (Excel Export)
# ─────────────────────────────────────────────────────────────────

def export_results(data_list: list, format: str = "excel") -> str | None:
    """
    İşlenen belge sonuçlarını ve finansal validasyon durumlarını
    profesyonel bir Excel dosyasına aktarır.
    
    Parametreler:
        data_list (list): Engine'den dönen ExtractedData objelerinin listesi
        format    (str) : Şimdilik sadece 'excel' destekleniyor
    
    Döndürdüğü Değer:
        str: Oluşturulan dosyanın yolu (başarısızsa None)
    """
    if format != "excel":
        print(f"[UYARI] '{format}' formatı desteklenmiyor. Sadece 'excel' kullanılabilir.")
        return None
    
    if not data_list:
        print("[BİLGİ] Dışa aktarılacak veri bulunamadı.")
        return None
    
    try:
        # Her belge için satır verisi oluştur
        satirlar = []
        for belge in data_list:
            # Finansal validasyonu çalıştır ve sonucunu al
            validasyon = validate_financials(belge)
            
            # Fatura detayları varsa doldur, yoksa boş bırak
            inv = getattr(belge, 'fatura_detaylari', None)
            
            satirlar.append({
                "Belge Türü"          : getattr(belge, 'document_type', 'Bilinmiyor'),
                "Özet"                : getattr(belge, 'document_summary', '')[:80],
                "Fatura No"           : getattr(inv, 'fatura_no', '-')    if inv else '-',
                "Tarih"               : getattr(inv, 'tarih', '-')        if inv else '-',
                "Satıcı"              : getattr(inv, 'satici_unvan', '-') if inv else '-',
                "Ara Toplam"          : getattr(inv, 'ara_toplam', '-')   if inv else '-',
                "KDV"                 : getattr(inv, 'kdv_tutari', '-')   if inv else '-',
                "Genel Toplam"        : getattr(inv, 'toplam_tutar', '-') if inv else '-',
                "Para Birimi"         : getattr(inv, 'para_birimi', '-')  if inv else '-',
                "AI Güven Skoru"      : f"%{int((getattr(inv, 'confidence_score', 0) or 0) * 100)}" if inv else '-',
                "Validasyon Durumu"   : "✅ Geçerli" if validasyon["valid"] else "❌ Hatalı",
                "Validasyon Detayı"   : validasyon["message"],
            })
        
        # Pandas DataFrame oluştur
        df = pd.DataFrame(satirlar)
        
        # Dosya adını tarih-saat damgasıyla oluştur: analiz_raporu_2026_04_11.xlsx
        tarih_damgasi = datetime.now().strftime("%Y_%m_%d")
        dosya_adi = f"analiz_raporu_{tarih_damgasi}.xlsx"
        
        # Çıktı dizinini oluştur (yoksa)
        cikti_dizini = Path("data/output")
        cikti_dizini.mkdir(parents=True, exist_ok=True)
        dosya_yolu = cikti_dizini / dosya_adi
        
        # Excel yazıcısını başlat
        with pd.ExcelWriter(dosya_yolu, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Analiz Raporu", index=False)
            
            ws = writer.sheets["Analiz Raporu"]
            
            # Başlık satırını stillendir: Koyu mavi arka plan + Beyaz kalın yazı
            baslik_doldurma = PatternFill("solid", fgColor="1F3864")
            baslik_yazi     = Font(color="FFFFFF", bold=True, size=11)
            ortalama        = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for hucre in ws[1]:  # 1. satır = başlık satırı
                hucre.fill      = baslik_doldurma
                hucre.font      = baslik_yazi
                hucre.alignment = ortalama
            
            # Satır yüksekliğini ve hücre hizalamasını ayarla
            ws.row_dimensions[1].height = 30
            for satir in ws.iter_rows(min_row=2):
                for hucre in satir:
                    hucre.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Sütun genişliklerini içeriğe göre otomatik ayarla
            for sutun_idx, sutun_hucreleri in enumerate(ws.columns, 1):
                max_genislik = 0
                for hucre in sutun_hucreleri:
                    if hucre.value:
                        # Her hücrenin metin uzunluğunu ölç, en uzunu seç
                        max_genislik = max(max_genislik, len(str(hucre.value)))
                # Minimum 12, maksimum 60 karakter genişliği uygula
                ws.column_dimensions[get_column_letter(sutun_idx)].width = min(max(max_genislik + 4, 12), 60)
        
        print(f"[BAŞARILI] Analiz raporu oluşturuldu → {dosya_yolu}")
        return str(dosya_yolu)
    
    except Exception as e:
        print(f"[HATA] Excel raporu oluşturulamadı: {e}")
        return None


# ─────────────────────────────────────────────────────────────────
# ANA AKIŞ
# ─────────────────────────────────────────────────────────────────

def main():
    """
    Sistemin dışarıdan tetiklendiği ilk çalıştırılabilir script.
    Engine → Validasyon → Excel Rapor sıralamasıyla çalışır.
    """
    app_engine = Engine()
    
    try:
        # Eski event loop yapısı yerine Python 3.7+ güvenli asenkron koşturucusu (asyncio.run) kullanılarak Engine tetiklenir.
        sonuclar = asyncio.run(app_engine.run())
        
        # Engine sonuç döndürüyorsa validasyon + raporlama adımlarını çalıştır
        if sonuclar:
            print("\n[BİLGİ] Finansal Çapraz Denetim (Cross-Check Validation) başlatılıyor...")
            for belge in sonuclar:
                validasyon = validate_financials(belge)
                print(f"  → {validasyon['message']}")
            
            print("\n[BİLGİ] Excel Raporu oluşturuluyor...")
            export_results(sonuclar, format="excel")
            
    except KeyboardInterrupt:
        # Kullanıcı 'Ctrl+C' yaparak sistemi durdurduğunda ortaya çıkan kaba hatayı gizleyip insancıl bir metin yazarız (Hata Yönetimi prensibi)
        print("\n[BİLGİ] İşlem kullanıcı tarafından elle (manuel) durduruldu.")
    except Exception as e:
        print(f"\n[KRİTİK HATA] Ana akış rutini (main.py) çalıştırılamadı → {e}")

if __name__ == "__main__":
    main()
