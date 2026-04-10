import os
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from src.schemas import ExtractedData
from src.config import settings

def export_data(data_list: List[ExtractedData], output_filename: str = "otonom_sistem_sonuclari.xlsx"):
    """
    Pandas yerine Openpyxl altyapısını kullanarak verileri görsel olarak zenginleştirir.
    Hücre renklendirme, koşullu biçimlendirme ve Hyperlink destekleri vardır.
    """
    output_dir = settings.get_output_path()
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / output_filename
    
    wb = Workbook()
    
    # 1. FATURALAR SAYFASI BAŞLIKLARI
    inv_sheet = wb.active
    inv_sheet.title = "Faturalar"
    
    inv_headers = [
        "Dosya Linki", "AI Durum Denetimi", "Fatura No", "Tarih", "Satıcı Ünvanı", "Alıcı Ünvanı", 
        "Ara Toplam", "KDV Tutarı", "Toplam Tutar", "Para Birimi", "Güven Skoru", "Kalemler"
    ]
    inv_sheet.append(inv_headers)
    
    # 2. SÖZLEŞMELER SAYFASI BAŞLIKLARI
    ctr_sheet = wb.create_sheet(title="Sözleşmeler")
    ctr_headers = [
        "Dosya Linki", "AI Durum Denetimi", "Sözleşme Konusu", "Tarih", "Taraflar", 
        "Geçerlilik Süresi", "Fesih Şartı Var Mı?", "Güven Skoru"
    ]
    ctr_sheet.append(ctr_headers)
    
    # Kurumsal Stil ve Renk Tanımları
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD") # Mavi Başlık
    
    error_fill = PatternFill("solid", fgColor="FFC7CE")      # Açık Kırmızı (KDV Hatası vb)
    warning_fill = PatternFill("solid", fgColor="FFEB9C")    # Turuncu/Sarı (Düşük Güven)
    duplicate_fill = PatternFill("solid", fgColor="E0E0E0")  # Gri (Çifte Kayıt)
    
    for sheet, headers in [(inv_sheet, inv_headers), (ctr_sheet, ctr_headers)]:
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(col_num)].width = 22

    # İstatistikler için Sayaçlar
    total_invoices = 0
    total_contracts = 0
    total_revenue = 0.0
    error_count = 0
    
    for row_idx, item in enumerate(data_list):
        is_invoice = item.fatura_detaylari is not None
        is_contract = item.sozlesme_detaylari is not None
        
        # Link Formülü (Hyperlink) oluştur
        file_path = item.source_file_path if item.source_file_path else ""
        link_formula = f'=HYPERLINK("{file_path}", "Dosyayı Aç")' if file_path else "Link Yok"
        
        # Hata Türüne Göre Satır Rengini Belirleme
        row_fill = None
        status_texts = []
        
        if getattr(item, 'is_duplicate', False):
            row_fill = duplicate_fill
            status_texts.append("Mükerrer")
        elif getattr(item, 'math_error', False):
            row_fill = error_fill
            status_texts.append("KDV Tutarsız")
        elif getattr(item, 'low_confidence', False):
            row_fill = warning_fill
            status_texts.append("Riskli Veri (Low Conf)")
            
        status_str = " + ".join(status_texts) if status_texts else "Kusursuz"
        
        if row_fill is not None:
             error_count += 1
             
        # Fatura Akışı
        if is_invoice:
            inv = item.fatura_detaylari
            
            satici = inv.satici_unvan if inv.satici_unvan else "DİKKAT: Eksik Veri"
            alici = inv.alici_unvan if inv.alici_unvan else "DİKKAT: Eksik Veri"
            
            row_data = [
                link_formula,
                status_str,
                inv.fatura_no,
                inv.tarih,
                satici,
                alici,
                inv.ara_toplam,
                inv.kdv_tutari,
                inv.toplam_tutar,
                inv.para_birimi,
                inv.confidence_score,
                ", ".join(inv.kalemler) if getattr(inv, 'kalemler', None) else ""
            ]
            
            inv_sheet.append(row_data)
            current_row = inv_sheet.max_row
            
            # Renklendirme uygula
            if row_fill:
                 for col_num in range(1, len(row_data) + 1):
                      inv_sheet.cell(row=current_row, column=col_num).fill = row_fill
                      
            # Sütunları standart muhasebe sayı biçimine dönüştür (Görsel hatayı -######- engellemek için)
            for col_num in [7, 8, 9]:
                 cell = inv_sheet.cell(row=current_row, column=col_num)
                 if isinstance(cell.value, (int, float)):
                      cell.number_format = '#,##0.00'
            
            # İstatistik artırımı
            total_invoices += 1
            if inv.toplam_tutar:
                 total_revenue += inv.toplam_tutar
                 
        # Sözleşme Akışı
        elif is_contract:
            ctr = item.sozlesme_detaylari
            row_data = [
                link_formula,
                status_str,
                ctr.sozlesme_konusu,
                ctr.sozlesme_tarihi,
                ", ".join(ctr.taraflar) if getattr(ctr, 'taraflar', None) else "",
                ctr.gecerlilik_suresi,
                "Mevcut" if ctr.fesih_sartlari_var_mi else "Yok",
                ctr.confidence_score
            ]
            ctr_sheet.append(row_data)
            current_row = ctr_sheet.max_row
            
            if row_fill:
                 for col_num in range(1, len(row_data) + 1):
                      ctr_sheet.cell(row=current_row, column=col_num).fill = row_fill

            total_contracts += 1

    # 3. ÖZET (SUMMARY) TABLO SAYFASI
    summary_sheet = wb.create_sheet(title="Özet İstatistikler")
    
    summary_sheet.append(["Metrik (KPI)", "Sistem Değeri"])
    summary_sheet.append(["İşlenen Başarılı Fatura Sayısı", total_invoices])
    summary_sheet.append(["İşlenen Başarılı Sözleşme Sayısı", total_contracts])
    summary_sheet.append(["Tanımlanan Toplam Ekonomik Hacim", total_revenue])
    summary_sheet.append(["Düşük Güven veya Hata Tespit Edilen Belge Sayısı", error_count])
    
    # Özet tablo stil ayarları
    for col_num, header in enumerate(["Metrik (KPI)", "Sistem Değeri"], 1):
        cell = summary_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="00B050") # Özet Yeşil
        summary_sheet.column_dimensions[get_column_letter(col_num)].width = 50
        
    summary_sheet.cell(row=4, column=2).number_format = '#,##0.00 "₺"'

    try:
        wb.save(str(report_path))
        print(f"[BAŞARILI] Mükemmelleştirilmiş Kurumsal Excel Raporu '{report_path}' üzerine kaydedildi.")
    except Exception as e:
        print(f"[HATA] Mükemmel Excel dosyası kaydedilirken klasör izni hatası oluştu: {e}")

def export_db_report(invoices: List, contracts: List, output_filename: str = "otonom_sistem_sonuclari.xlsx"):
    """
    Veritabanında kalıcı olarak kayıtlı olan ve onaylanmış kayıtları (InvoiceRecord ve ContractRecord)
    okuyarak Tam Senkronize Openpyxl raporunu oluşturur.
    """
    output_dir = settings.get_output_path()
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / output_filename
    
    wb = Workbook()
    
    # 1. FATURALAR SAYFASI
    inv_sheet = wb.active
    inv_sheet.title = "Faturalar"
    
    inv_headers = [
        "Dosya Linki", "AI Durum Denetimi", "Fatura No", "Tarih", "Satıcı Ünvanı", "Alıcı Ünvanı", 
        "Ara Toplam", "KDV Tutarı", "Toplam Tutar", "Para Birimi", "Güven Skoru", "Kalemler"
    ]
    inv_sheet.append(inv_headers)
    
    # 2. SÖZLEŞMELER SAYFASI
    ctr_sheet = wb.create_sheet(title="Sözleşmeler")
    ctr_headers = [
        "Dosya Linki", "AI Durum Denetimi", "Sözleşme Konusu", "Tarih", "Taraflar", 
        "Geçerlilik Süresi", "Fesih Şartı Var Mı?", "Güven Skoru"
    ]
    ctr_sheet.append(ctr_headers)
    
    # Kurumsal Stil ve Renk Tanımları
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD") # Mavi Başlık
    error_fill = PatternFill("solid", fgColor="FFC7CE")      # Açık Kırmızı (KDV Hatası vb)
    warning_fill = PatternFill("solid", fgColor="FFEB9C")    # Turuncu/Sarı (Düşük Güven)
    
    for sheet, headers in [(inv_sheet, inv_headers), (ctr_sheet, ctr_headers)]:
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(col_num)].width = 22

    # İstatistikler için Sayaçlar
    total_invoices = 0
    total_contracts = 0
    total_revenue = 0.0
    error_count = 0
    
    # FATURALAR DÖNGÜSÜ
    for inv in invoices:
        file_path = inv.source_file if inv.source_file else ""
        link_formula = f'=HYPERLINK("{file_path}", "Dosyayı Aç")' if file_path else "Link Yok"
        
        row_fill = None
        status_texts = []
        
        if inv.math_error:
            row_fill = error_fill
            status_texts.append("KDV Tutarsız")
        elif inv.low_confidence:
            row_fill = warning_fill
            status_texts.append("Riskli Veri (Low Conf)")
            
        status_str = " + ".join(status_texts) if status_texts else "Kusursuz"
        
        if row_fill is not None:
             error_count += 1
             
        satici = inv.satici_unvan if inv.satici_unvan else "DİKKAT: Eksik Veri"
        alici = inv.alici_unvan if inv.alici_unvan else "DİKKAT: Eksik Veri"
        
        row_data = [
            link_formula,
            status_str,
            inv.fatura_no,
            inv.tarih,
            satici,
            alici,
            inv.ara_toplam,
            inv.kdv_tutari,
            inv.toplam_tutar,
            inv.para_birimi,
            inv.confidence_score,
            inv.kalemler if inv.kalemler else ""
        ]
        
        inv_sheet.append(row_data)
        current_row = inv_sheet.max_row
        
        # Renklendirme uygula
        if row_fill:
             for col_num in range(1, len(row_data) + 1):
                  inv_sheet.cell(row=current_row, column=col_num).fill = row_fill
                  
        # Para Birimi Sembolleri Format Çakışmasını Önleme
        for col_num in [7, 8, 9]:
             cell = inv_sheet.cell(row=current_row, column=col_num)
             if isinstance(cell.value, (int, float)):
                  cell.number_format = '#,##0.00'
        
        total_invoices += 1
        if inv.toplam_tutar:
             total_revenue += inv.toplam_tutar

    # SÖZLEŞMELER DÖNGÜSÜ
    for ctr in contracts:
        file_path = ctr.source_file if ctr.source_file else ""
        link_formula = f'=HYPERLINK("{file_path}", "Dosyayı Aç")' if file_path else "Link Yok"
        
        row_fill = None
        status_texts = []
        
        if ctr.low_confidence:
            row_fill = warning_fill
            status_texts.append("Riskli Veri (Low Conf)")
            
        status_str = " + ".join(status_texts) if status_texts else "Kusursuz"
        
        if row_fill is not None:
             error_count += 1

        row_data = [
            link_formula,
            status_str,
            ctr.sozlesme_konusu,
            ctr.sozlesme_tarihi,
            ctr.taraflar if ctr.taraflar else "",
            ctr.gecerlilik_suresi,
            "Mevcut" if ctr.fesih_sartlari_var_mi else "Yok",
            ctr.confidence_score
        ]
        ctr_sheet.append(row_data)
        current_row = ctr_sheet.max_row
        
        if row_fill:
             for col_num in range(1, len(row_data) + 1):
                  ctr_sheet.cell(row=current_row, column=col_num).fill = row_fill

        total_contracts += 1

    # 3. ÖZET (SUMMARY) TABLO SAYFASI
    summary_sheet = wb.create_sheet(title="Özet İstatistikler")
    summary_sheet.append(["Metrik (KPI)", "Sistem Değeri"])
    summary_sheet.append(["Sistemdeki Orijinal Fatura Sayısı", total_invoices])
    summary_sheet.append(["Sistemdeki Orijinal Sözleşme Sayısı", total_contracts])
    summary_sheet.append(["Tanımlanan Toplam Ekonomik Hacim", total_revenue])
    summary_sheet.append(["Düşük Güven veya Hata Tespit Edilen Belge Sayısı", error_count])
    
    for col_num, header in enumerate(["Metrik (KPI)", "Sistem Değeri"], 1):
        cell = summary_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="00B050") # Özet Yeşil
        summary_sheet.column_dimensions[get_column_letter(col_num)].width = 50
        
    summary_sheet.cell(row=4, column=2).number_format = '#,##0.00 "₺"'

    try:
        wb.save(str(report_path))
        print(f"[BAŞARILI] Tam Senkron Kurumsal Excel Raporu '{report_path}' oluşturuldu.")
    except Exception as e:
        print(f"[HATA] Excel dosyası kaydedilirken hata oluştu: {e}")
